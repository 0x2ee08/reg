import os
import openpyxl
from openpyxl.descriptors.base import String

def layDL(path, cellname):
    wb = openpyxl.load_workbook(path)
    Sheet1 = wb['Sheet']
    wb.close()
    return Sheet1[cellname].value

def updateDL(path, cellTHS, cellLHS, cellDD, cellTN, cellKT,valueTHS, valueLHS, valueDD, valueTN, valueKT):
    try:
        wb = openpyxl.load_workbook(path)
    except:
        wb=openpyxl.Workbook()
        wb.save(path)
        wb = openpyxl.load_workbook(path)

    Sheet1 = wb['Sheet']

    Sheet1[cellTHS].value = valueTHS
    Sheet1[cellLHS].value = valueLHS
    Sheet1[cellDD].value = valueDD
    Sheet1[cellTN].value = valueTN
    Sheet1[cellKT].value = valueKT

    wb.close()
    wb.save(path)
